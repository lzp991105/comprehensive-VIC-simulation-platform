import subprocess
import sys
import codecs
import os
import shutil
import io
import time
import json
import threading
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.Qt import QWidget, QMainWindow
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from pysnooperDB import DataBaseFunc as DB
from subprocess import run
from PIL import Image
from PIL import ImageTk  # 如果你计划在GUI应用中使用图像，比如Tkinter
from openpyxl import Workbook, load_workbook
import pandas as pd
import random
import socket
import tkinter as tk
from tkinter import filedialog
import math
from Al import *
import images


# global count
count = 0
# global selected_text
selected_road_text = 0
selected_comm_text = 0

Pcom_v = 0.8
Ptran_b = 5  # W
Ptran_c = 5
Ptran_v = 1  # W
h_base = 5
h_cloud = 2
Sita = 0.1

count_InitTask = 0

# 在模块层次上声明全局变量
external_car_fre = 400
external_base_x = 600
external_base_y = 500
external_base_band = 10
external_base_fre = 8
external_base_cov = 100

external_cloud_x = 0
external_cloud_y = 0
external_cloud_band = 20
external_cloud_fre = 20
external_cloud_cov = 20

task_loss = 0
SC = 0

# 这段代码是使用 PyQt5 库创建的图形用户界面 (GUI)。它定义了一个名为 Ui_controlbutton 的类，该类中的 setupUi 方法用于设置和初始化 GUI 的各个组件
class Ui_controlbutton(object):
    def setupUi(self, controlbutton):
        # 设置窗口的名称和大小
        controlbutton.setObjectName("controlbutton")
        controlbutton.resize(300, 300)
        # 创建一个水平布局，并将其添加到 controlbutton 窗口。
        self.horizontalLayout_4 = QtWidgets.QHBoxLayout(controlbutton)
        self.horizontalLayout_4.setObjectName("horizontalLayout_4")
        # 创建一个垂直布局
        self.verticalLayout_2 = QtWidgets.QVBoxLayout()
        self.verticalLayout_2.setObjectName("verticalLayout_2")

        self.horizontalLayout_8 = QtWidgets.QHBoxLayout()
        self.verticalLayout_2.addLayout(self.horizontalLayout_8)
        self.horizontalLayout_4.setObjectName("horizontalLayout_8")

        # 车辆计算频率
        self.vehicle_frequency = QtWidgets.QLabel(controlbutton)
        self.vehicle_frequency.setObjectName("vehicle_frequency")
        self.horizontalLayout_8.addWidget(self.vehicle_frequency)

        # 文本框
        self.vfrequency = QtWidgets.QLineEdit(controlbutton)
        self.vfrequency.setEnabled(True)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.vfrequency.sizePolicy().hasHeightForWidth())
        self.vfrequency.setSizePolicy(sizePolicy)
        self.vfrequency.setAlignment(QtCore.Qt.AlignCenter)
        self.vfrequency.setObjectName("vfrequency")
        self.horizontalLayout_8.addWidget(self.vfrequency)
        self.vfrequency.editingFinished.connect(self.vfrequency_input_finished)

        # 频率单位
        self.label_frequency = QtWidgets.QLabel(controlbutton)
        self.label_frequency.setObjectName("label_frequency")
        self.horizontalLayout_8.addWidget(self.label_frequency)

        self.verticalLayout_2.addLayout(self.horizontalLayout_8)

        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")

        # 创建一个标签，并将其添加到 controlbutton 窗口
        # 基站
        self.label_base = QtWidgets.QLabel(controlbutton)
        self.label_base.setObjectName("base1")
        self.horizontalLayout.addWidget(self.label_base)

        self.label_base_x = QtWidgets.QLabel(controlbutton)
        self.label_base_x.setObjectName("label_base_x")
        self.horizontalLayout.addWidget(self.label_base_x)

        # 文本框
        self.base_x = QtWidgets.QLineEdit(controlbutton)
        self.base_x.setEnabled(True)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.base_x.sizePolicy().hasHeightForWidth())
        self.base_x.setSizePolicy(sizePolicy)
        self.base_x.setAlignment(QtCore.Qt.AlignCenter)
        self.base_x.setObjectName("location1_x")
        self.horizontalLayout.addWidget(self.base_x)
        self.base_x.editingFinished.connect(self.bx_input_finished)

        self.label_base_y = QtWidgets.QLabel(controlbutton)
        self.label_base_y.setObjectName("label_base_y")
        self.horizontalLayout.addWidget(self.label_base_y)

        spacerItem = QtWidgets.QSpacerItem(80, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout.addItem(spacerItem)

        self.base_y = QtWidgets.QLineEdit(controlbutton)
        self.base_y.setEnabled(True)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.base_y.sizePolicy().hasHeightForWidth())
        self.base_y.setSizePolicy(sizePolicy)
        self.base_y.setAlignment(QtCore.Qt.AlignCenter)
        self.base_y.setObjectName("base_y")
        self.horizontalLayout.addWidget(self.base_y)
        self.base_y.editingFinished.connect(self.by_input_finished)

        # spacerItem = QtWidgets.QSpacerItem(80, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        # self.horizontalLayout.addItem(spacerItem)

        # 将水平布局添加到垂直布局。
        self.verticalLayout_2.addLayout(self.horizontalLayout)
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")

        # 带宽大小
        self.label_base_bandwidth = QtWidgets.QLabel(controlbutton)
        self.label_base_bandwidth.setObjectName("label_base_bandwidth")
        self.horizontalLayout_2.addWidget(self.label_base_bandwidth)
        # 带宽大小输入框
        self.base_bandwidth = QtWidgets.QLineEdit(controlbutton)
        self.base_bandwidth.setEnabled(True)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.base_bandwidth.sizePolicy().hasHeightForWidth())
        self.base_bandwidth.setSizePolicy(sizePolicy)
        self.base_bandwidth.setAlignment(QtCore.Qt.AlignCenter)
        self.base_bandwidth.setObjectName("base_bandwidth")
        self.horizontalLayout_2.addWidget(self.base_bandwidth)
        self.base_bandwidth.editingFinished.connect(self.bbandwidth_input_finished)

        # 带宽大小单位
        self.label_Hz = QtWidgets.QLabel(controlbutton)
        self.label_Hz.setObjectName("label_Hz")
        self.horizontalLayout_2.addWidget(self.label_Hz)
        self.verticalLayout_2.addLayout(self.horizontalLayout_2)

        # 计算频率
        self.label_base_Fre = QtWidgets.QLabel(controlbutton)
        self.label_base_Fre.setObjectName("label_Fre")
        self.horizontalLayout_2.addWidget(self.label_base_Fre)
        # 计算频率输入框
        self.base_frequency = QtWidgets.QLineEdit(controlbutton)
        self.base_frequency.setEnabled(True)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.base_frequency.sizePolicy().hasHeightForWidth())
        self.base_frequency.setSizePolicy(sizePolicy)
        self.base_frequency.setAlignment(QtCore.Qt.AlignCenter)
        self.base_frequency.setObjectName("base_frequency")
        self.horizontalLayout_2.addWidget(self.base_frequency)
        self.base_frequency.editingFinished.connect(self.bfrequency_input_finished)

        # 计算频率单位
        self.label_M = QtWidgets.QLabel(controlbutton)
        self.label_M.setObjectName("label_M")
        self.horizontalLayout_2.addWidget(self.label_M)



        # 将水平布局添加到垂直布局。
        # self.verticalLayout_2.addLayout(self.horizontalLayout)
        self.horizontalLayout_7 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_7.setObjectName("horizontalLayout_7")
        # 覆盖范围
        self.label_base_Cov = QtWidgets.QLabel(controlbutton)
        self.label_base_Cov.setObjectName("label_base_Cov")
        self.horizontalLayout_7.addWidget(self.label_base_Cov)
        # 覆盖范围文本框
        self.base_coverage = QtWidgets.QLineEdit(controlbutton)
        self.base_coverage.setEnabled(False)
        self.base_coverage.setAlignment(QtCore.Qt.AlignCenter)
        self.base_coverage.setObjectName("base_coverage")
        self.horizontalLayout_7.addWidget(self.base_coverage)

        # 覆盖范围单位
        self.label_base_Meters = QtWidgets.QLabel(controlbutton)
        self.label_base_Meters.setObjectName("label_base_Meters")
        self.horizontalLayout_7.addWidget(self.label_base_Meters)
        self.verticalLayout_2.addLayout(self.horizontalLayout_7)

        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.verticalLayout_2.addLayout(self.horizontalLayout_3)
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")

        # 云中心
        self.label_cloud = QtWidgets.QLabel(controlbutton)
        self.label_cloud.setObjectName("cloud")
        self.horizontalLayout_3.addWidget(self.label_cloud)

        self.label_cloud_x = QtWidgets.QLabel(controlbutton)
        self.label_cloud_x.setObjectName("cloud_x")
        self.horizontalLayout_3.addWidget(self.label_cloud_x)

        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_3.addItem(spacerItem)
        self.cloud_x = QtWidgets.QLineEdit(controlbutton)
        self.cloud_x.setEnabled(True)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.cloud_x.sizePolicy().hasHeightForWidth())
        self.cloud_x.setSizePolicy(sizePolicy)
        self.cloud_x.setAlignment(QtCore.Qt.AlignCenter)
        self.cloud_x.setObjectName("location2_x")
        self.horizontalLayout_3.addWidget(self.cloud_x)
        self.cloud_x.editingFinished.connect(self.cx_input_finished)

        self.label_cloud_y = QtWidgets.QLabel(controlbutton)
        self.label_cloud_y.setObjectName("base2_y")
        self.horizontalLayout_3.addWidget(self.label_cloud_y)
        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_3.addItem(spacerItem)

        self.cloud_y = QtWidgets.QLineEdit(controlbutton)
        self.cloud_y.setEnabled(True)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.cloud_y.sizePolicy().hasHeightForWidth())
        self.cloud_y.setSizePolicy(sizePolicy)
        self.cloud_y.setAlignment(QtCore.Qt.AlignCenter)
        self.cloud_y.setObjectName("location2_y")
        self.horizontalLayout_3.addWidget(self.cloud_y)
        self.cloud_y.editingFinished.connect(self.cy_input_finished)

        spacerItem1 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout.addItem(spacerItem1)

        # 将水平布局添加到垂直布局。
        self.verticalLayout_2.addLayout(self.horizontalLayout)
        self.horizontalLayout_6 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_6.setObjectName("horizontalLayout_2")

        # 带宽大小
        self.label_cloud_bandwidth = QtWidgets.QLabel(controlbutton)
        self.label_cloud_bandwidth.setObjectName("label_22")
        self.horizontalLayout_6.addWidget(self.label_cloud_bandwidth)
        # 带宽大小输入框
        self.cloud_bandwidth = QtWidgets.QLineEdit(controlbutton)
        self.cloud_bandwidth.setEnabled(True)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.cloud_bandwidth.sizePolicy().hasHeightForWidth())
        self.cloud_bandwidth.setSizePolicy(sizePolicy)
        self.cloud_bandwidth.setAlignment(QtCore.Qt.AlignCenter)
        self.cloud_bandwidth.setObjectName("bandwidth2")
        self.horizontalLayout_6.addWidget(self.cloud_bandwidth)
        self.cloud_bandwidth.editingFinished.connect(self.cbandwidth_input_finished)

        # 带宽大小单位
        self.label_Hz2 = QtWidgets.QLabel(controlbutton)
        self.label_Hz2.setObjectName("label__Hz2")
        self.horizontalLayout_6.addWidget(self.label_Hz2)

        self.verticalLayout_2.addLayout(self.horizontalLayout_6)

        # 计算频率
        self.label_cloud_Fre = QtWidgets.QLabel(controlbutton)
        self.label_cloud_Fre.setObjectName("label_cloud_Fre")
        self.horizontalLayout_6.addWidget(self.label_cloud_Fre)
        # 计算频率输入框
        self.cloud_frequency = QtWidgets.QLineEdit(controlbutton)
        self.cloud_frequency.setEnabled(True)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.cloud_frequency.sizePolicy().hasHeightForWidth())
        self.cloud_frequency.setSizePolicy(sizePolicy)
        self.cloud_frequency.setAlignment(QtCore.Qt.AlignCenter)
        self.cloud_frequency.setObjectName("frequency")
        self.horizontalLayout_6.addWidget(self.cloud_frequency)
        self.cloud_frequency.editingFinished.connect(self.cfrequency_input_finished)

        # 计算频率单位
        self.label_M2 = QtWidgets.QLabel(controlbutton)
        self.label_M2.setObjectName("label__M2")
        self.horizontalLayout_6.addWidget(self.label_M2)

        # 将水平布局添加到垂直布局。
        # self.verticalLayout_2.addLayout(self.horizontalLayout)
        self.horizontalLayout_8 = QtWidgets.QHBoxLayout()
        self.verticalLayout_2.addLayout(self.horizontalLayout_8)

        self.horizontalLayout_8.setObjectName("horizontalLayout_8")
        # 覆盖范围
        self.label_cloud_Cov = QtWidgets.QLabel(controlbutton)
        self.label_cloud_Cov.setObjectName("label_cloud_Cov")
        self.horizontalLayout_8.addWidget(self.label_cloud_Cov)
        # 覆盖范围文本框
        self.cloud_coverage = QtWidgets.QLineEdit(controlbutton)
        self.cloud_coverage.setEnabled(False)
        self.cloud_coverage.setAlignment(QtCore.Qt.AlignCenter)
        self.cloud_coverage.setObjectName("coverage")
        self.horizontalLayout_8.addWidget(self.cloud_coverage)

        # 覆盖范围单位
        self.label_cloud_Meters = QtWidgets.QLabel(controlbutton)
        self.label_cloud_Meters.setObjectName("label_cloud_Meters")
        self.horizontalLayout_8.addWidget(self.label_cloud_Meters)

        spacerItem5 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_3.addItem(spacerItem5)
        self.horizontalLayout_3.setStretch(0, 1)
        self.horizontalLayout_3.setStretch(1, 1)
        self.horizontalLayout_3.setStretch(2, 1)
        self.horizontalLayout_3.setStretch(3, 1)
        self.horizontalLayout_3.setStretch(4, 2)
        self.horizontalLayout_3.setStretch(5, 1)

        self.verticalLayout_2.setStretch(0, 1)
        self.verticalLayout_2.setStretch(1, 1)
        self.verticalLayout_2.setStretch(2, 1)
        self.verticalLayout_2.setStretch(3, 1)
        self.verticalLayout_2.setStretch(4, 1)
        self.verticalLayout_2.setStretch(5, 1)
        # 将垂直布局添加到另一个水平布局。
        self.horizontalLayout_4.addLayout(self.verticalLayout_2)
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")

        self.run = QtWidgets.QPushButton(controlbutton)
        self.run.setMinimumSize(QtCore.QSize(0, 50))
        self.run.setObjectName("run")
        self.verticalLayout.addWidget(self.run)
        self.run.clicked.connect(self.InitResources)

        self.horizontalLayout_4.addLayout(self.verticalLayout)
        self.horizontalLayout_4.setStretch(0, 3)
        self.horizontalLayout_4.setStretch(1, 1)
        # 调用 retranslateUi 方法，该方法通常用于设置或更新组件的文本。
        self.retranslateUi(controlbutton)

    def vfrequency_input_finished(self):
        user_input = self.vfrequency.text()
        try:
            input_number = float(user_input)
            if input_number < 0:
                QtWidgets.QMessageBox.information(None, "提示", "频率不能小于0")
                self.vfrequency.clear()  # 清除文本框内容
            # else:
            #     print(2)
        except ValueError:
            QtWidgets.QMessageBox.information(None, "错误", "请输入数字")
            self.vfrequency.clear()  # 清除文本框内容

    def bx_input_finished(self):
        user_input = self.base_x.text()
        try:
            input_number = float(user_input)
            if input_number < 0:
                QtWidgets.QMessageBox.information(None, "提示", "横坐标不能小于0")
                self.base_x.clear()  # 清除文本框内容
        except ValueError:
            QtWidgets.QMessageBox.information(None, "错误", "请输入数字")
            self.base_x.clear()  # 清除文本框内容

    def by_input_finished(self):
        user_input = self.base_y.text()
        try:
            input_number = float(user_input)
            if input_number < 0:
                QtWidgets.QMessageBox.information(None, "提示", "纵坐标不能小于0")
                self.base_y.clear()  # 清除文本框内容
        except ValueError:
            QtWidgets.QMessageBox.information(None, "错误", "请输入数字")
            self.base_y.clear()  # 清除文本框内容

    def bbandwidth_input_finished(self):
        user_input = self.base_bandwidth.text()
        try:
            input_number = float(user_input)
            if input_number < 0:
                QtWidgets.QMessageBox.information(None, "提示", "带宽不能小于0")
                self.base_bandwidth.clear()  # 清除文本框内容
        except ValueError:
            QtWidgets.QMessageBox.information(None, "错误", "请输入数字")
            self.base_bandwidth.clear()  # 清除文本框内容

    def bfrequency_input_finished(self):
        user_input = self.base_frequency.text()
        try:
            input_number = float(user_input)
            if input_number < 0:
                QtWidgets.QMessageBox.information(None, "提示", "频率不能小于0")
                self.base_frequency.clear()  # 清除文本框内容
        except ValueError:
            QtWidgets.QMessageBox.information(None, "错误", "请输入数字")
            self.base_frequency.clear()  # 清除文本框内容

    def cx_input_finished(self):
        user_input = self.cloud_x.text()
        try:
            input_number = float(user_input)
            if input_number < 0:
                QtWidgets.QMessageBox.information(None, "提示", "横坐标不能小于0")
                self.cloud_x.clear()  # 清除文本框内容

        except ValueError:
            QtWidgets.QMessageBox.information(None, "错误", "请输入数字")
            self.cloud_x.clear()  # 清除文本框内容

    def cy_input_finished(self):
        user_input = self.cloud_y.text()
        try:
            input_number = float(user_input)
            if input_number < 0:
                QtWidgets.QMessageBox.information(None, "提示", "纵坐标不能小于0")
                self.cloud_y.clear()  # 清除文本框内容
        except ValueError:
            QtWidgets.QMessageBox.information(None, "错误", "请输入数字")
            self.cloud_y.clear()  # 清除文本框内容

    def cbandwidth_input_finished(self):
        user_input = self.cloud_bandwidth.text()
        try:
            input_number = float(user_input)
            if input_number < 0:
                QtWidgets.QMessageBox.information(None, "提示", "带宽不能小于0")
                self.cloud_bandwidth.clear()  # 清除文本框内容
        except ValueError:
            QtWidgets.QMessageBox.information(None, "错误", "请输入数字")
            self.cloud_bandwidth.clear()  # 清除文本框内容

    def cfrequency_input_finished(self):
        user_input = self.cloud_frequency.text()
        try:
            input_number = float(user_input)
            if input_number < 0:
                QtWidgets.QMessageBox.information(None, "提示", "频率不能小于0")
                self.cloud_frequency.clear()  # 清除文本框内容
        except ValueError:
            QtWidgets.QMessageBox.information(None, "错误", "请输入数字")
            self.cloud_frequency.clear()  # 清除文本框内容

    def InitResources(self):
        current_vehicle_fre = self.vfrequency.text().strip()

        current_base_x = self.base_x.text().strip()
        # print(current_base_x)
        current_base_y = self.base_y.text().strip()
        current_base_bandwidth = self.base_bandwidth.text().strip()
        current_base_frequency = self.base_frequency.text().strip()

        current_cloud_x = self.cloud_x.text().strip()
        current_cloud_y = self.cloud_y.text().strip()
        current_cloud_bandwidth = self.cloud_bandwidth.text().strip()
        current_cloud_frequency = self.cloud_frequency.text().strip()

        global external_car_fre
        external_car_fre = float(current_vehicle_fre)
        global external_base_x
        external_base_x = float(current_base_x)
        global external_base_y
        external_base_y = float(current_base_y)
        global external_base_band
        external_base_band = float(current_base_bandwidth) * 1000
        global external_base_fre
        external_base_fre =float(current_base_frequency) * 1000
        global external_cloud_x
        external_cloud_x = float(current_cloud_x)
        global external_cloud_y
        external_cloud_y = float(current_cloud_y)
        global external_cloud_band
        external_cloud_band = float(current_cloud_bandwidth) * 1000
        global external_cloud_fre
        external_cloud_fre = float(current_cloud_frequency) * 1000

        if current_vehicle_fre and current_base_x and current_base_y and current_base_bandwidth and \
                current_base_frequency and current_cloud_x and current_cloud_y \
                and current_cloud_bandwidth and current_cloud_frequency:
            if count_InitTask > 0:
                desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
                excel_file_path = os.path.join(desktop_path, 'Dataset.xlsx')
                # excel_file_path = os.path.join(desktop_path, 'UNCON.xlsx')
                if os.path.exists(excel_file_path):
                    df = pd.read_excel(excel_file_path)
                    car_list = []
                    for index, row in df.iterrows():
                        row_list = list(row)
                        row_list[10] = row_list[8] / float(current_vehicle_fre)
                        row_list[11] = row_list[10] * Pcom_v

                        Tup_base = (row_list[9] * 1024 * 8) / (
                                    row_list[6] * 1000 * math.log2(1 + (Ptran_v * h_base / (Sita * Sita))))
                        Tdown_base = (row_list[9] * 1024 * 8 * 0.1) / (
                                    row_list[6] * 1000 * math.log2(1 + (Ptran_b * h_base / (Sita * Sita))))
                        Tcom_base = row_list[8] / row_list[7]
                        row_list[12] = Tup_base + Tdown_base + Tcom_base
                        row_list[13] = Tup_base * Ptran_v

                        Tup_cloud = (row_list[9] * 1024 * 8) / (
                                row_list[6] * 1000 * math.log2(1 + (Ptran_v * h_cloud / (Sita * Sita))))
                        Tdown_cloud = (row_list[9] * 1024 * 8 * 0.1) / (
                                row_list[6] * 1000 * math.log2(1 + (Ptran_c * h_cloud / (Sita * Sita))))
                        Tcom_cloud = row_list[8] / row_list[7]
                        row_list[14] = Tup_cloud + Tdown_cloud + Tcom_cloud
                        row_list[15] = Tup_cloud * Ptran_v
                        car_list.append(row_list)
                    # 创建Excel表格
                    wb = Workbook()
                    ws = wb.active

                    ws.append(
                        ['编号', '横坐标', '纵坐标', '速度', '时刻', '是否为任务车辆', '任务带宽需求', '任务频率需求',
                         '工作负载', '任务大小', '本地时间', '本地能耗', '基站时间', '基站能耗',
                         '云卸载时间','云卸载能耗'])
                    for row_data in car_list:
                        ws.append(row_data)
                    # 保存表格
                    wb.save(excel_file_path)
                    QtWidgets.QMessageBox.information(None, "提示", "生成成功")
                else:
                    QtWidgets.QMessageBox.information(None, "提示", "请先生成数据集")
            else:
                QtWidgets.QMessageBox.information(None, "提示", "请先进行任务设置")
        else:
            QtWidgets.QMessageBox.information(None, "提示", "请先进行边缘资源设置")

    def retranslateUi(self, controlbutton):
        _translate = QtCore.QCoreApplication.translate
        # 设置窗口标题为 "车辆边缘计算仿真平台"
        controlbutton.setWindowTitle(_translate("controlbutton", "仿真平台"))

        self.vehicle_frequency.setText(_translate("controlbutton", "车辆计算频率：   "))
        self.label_frequency.setText(_translate("controlbutton", "MHz"))

        self.label_base.setText(_translate("controlbutton", "基站信息       "))
        self.label_base_x.setText(_translate("controlbutton", "横坐标："))
        self.base_x.setText(_translate("controlbutton", "0"))
        self.label_base_y.setText(_translate("controlbutton", "纵坐标："))
        self.base_y.setText(_translate("controlbutton", "0"))
        self.label_base_bandwidth.setText(_translate("controlbutton", "带宽大小："))
        self.base_bandwidth.setText(_translate("MainWindow", "10"))
        self.label_Hz.setText(_translate("controlbutton", "MHz        "))
        self.label_base_Fre.setText(_translate("controlbutton", "计算频率："))
        self.base_frequency.setText(_translate("controlbutton", "8"))
        self.label_M.setText(_translate("controlbutton", "GHz"))
        self.label_base_Cov.setText(_translate("controlbutton", "覆盖范围："))
        self.base_coverage.setText(_translate("controlbutton", "150"))
        self.label_base_Meters.setText(_translate("controlbutton", "m"))

        self.label_cloud.setText(_translate("controlbutton", "云信息       "))
        self.label_cloud_x.setText(_translate("controlbutton", "横坐标："))
        self.cloud_x.setText(_translate("controlbutton", "0"))
        self.label_cloud_y.setText(_translate("controlbutton", "纵坐标："))
        self.cloud_y.setText(_translate("controlbutton", "0"))
        self.label_cloud_bandwidth.setText(_translate("controlbutton", "带宽大小："))
        self.cloud_bandwidth.setText(_translate("controlbutton", "20"))
        self.label_Hz2.setText(_translate("controlbutton", "MHz        "))
        self.label_cloud_Fre.setText(_translate("controlbutton", "计算频率："))
        self.cloud_frequency.setText(_translate("controlbutton", "20"))
        self.label_M2.setText(_translate("controlbutton", "GHz"))
        self.label_cloud_Cov.setText(_translate("controlbutton", "覆盖范围："))
        self.cloud_coverage.setText(_translate("controlbutton", "20"))
        self.label_cloud_Meters.setText(_translate("controlbutton", "Km"))
        self.run.setText(_translate("controlbutton", "初试化数据集"))


# 类名为controlbutton，它继承了QWidget和Ui_controlbutton两个父类。
class controlbutton(QWidget, Ui_controlbutton):
    # __init__是一个特殊的方法，称为类的构造函数。当创建这个类的新实例时，这个方法会被自动调用。
    def __init__(self):
        # 是调用父类的构造函数。在这里，它调用了QWidget和Ui_controlbutton的构造函数
        super(controlbutton, self).__init__()
        # 调用Ui_controlbutton类中定义的setupUi方法，用于设置用户界面。这个方法需要一个参数，这里传入的是self，表示的是当前类的实例。
        self.setupUi(self)


class Ui_controlbutton1(object):
    def setupUi(self, controlbutton1):
        # 设置窗口的名称和大小
        controlbutton1.setObjectName("controlbutton1")
        controlbutton1.resize(300, 300)

        # 创建一个垂直布局
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")

        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")

        # 创建一个标签，并将其添加到 controlbutton 窗口
        # 带宽
        self.bandwidth_choose = QtWidgets.QLabel(controlbutton1)
        self.bandwidth_choose.setObjectName("bandwidth_choose")
        self.horizontalLayout.addWidget(self.bandwidth_choose)

        self.comboBox = QtWidgets.QComboBox(controlbutton1)  # 创建一个下拉文本框
        self.comboBox.setObjectName("comboBox")  # 设置对象名称
        self.comboBox.addItem("250,300,350")  # 添加选项
        # self.comboBox.addItem("500,600,700")  # 添加选项
        self.comboBox.addItem("360,380,400")  # 添加选项
        self.comboBox.addItem("250,350,360,380,400")  # 添加选项
        # 将下拉文本框添加到水平布局
        self.horizontalLayout.addWidget(self.comboBox)

        # 带宽大小单位
        self.label_Hz = QtWidgets.QLabel(controlbutton1)
        self.label_Hz.setObjectName("label__Hz")
        self.horizontalLayout.addWidget(self.label_Hz)

        # 将水平布局添加到垂直布局。
        self.verticalLayout.addLayout(self.horizontalLayout)

        self.horizontalLayout_5 = QtWidgets.QHBoxLayout()
        self.verticalLayout.addLayout(self.horizontalLayout_5)
        self.horizontalLayout_5.setObjectName("horizontalLayout_5")

        # 频率大小
        self.fre_need = QtWidgets.QLabel(controlbutton1)
        self.fre_need.setObjectName("fre_need")
        self.horizontalLayout_5.addWidget(self.fre_need)

        self.comboBox4 = QtWidgets.QComboBox(controlbutton1)  # 创建一个下拉文本框
        self.comboBox4.setObjectName("comboBox3")  # 设置对象名称
        # self.comboBox4.addItem("400,540,600")  # 添加选项
        self.comboBox4.addItem("600,800,1000")  # 添加选项
        self.comboBox4.addItem("450,540,600")  # 添加选项
        self.comboBox4.addItem("400,500,530,570,600")  # 添加选项
        # 将下拉文本框添加到水平布局
        self.horizontalLayout_5.addWidget(self.comboBox4)

        # 频率单位
        self.label_fn = QtWidgets.QLabel(controlbutton1)
        self.label_fn.setObjectName("label_fn")
        self.horizontalLayout_5.addWidget(self.label_fn)

        self.verticalLayout.addLayout(self.horizontalLayout_5)

        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")

        # 工作负载大小
        self.workload = QtWidgets.QLabel(controlbutton1)
        self.workload.setObjectName("workload")
        self.horizontalLayout_2.addWidget(self.workload)

        self.comboBox2 = QtWidgets.QComboBox(controlbutton1)  # 创建一个下拉文本框
        self.comboBox2.setObjectName("comboBox")  # 设置对象名称
        self.comboBox2.addItem("9,50,98")  # 添加选项
        self.comboBox2.addItem("27,150,300")  # 添加选项
        self.comboBox2.addItem("9,27,50,98,150,300")  # 添加选项
        # 将下拉文本框添加到水平布局
        self.horizontalLayout_2.addWidget(self.comboBox2)

        # 工作负载单位
        self.label_M = QtWidgets.QLabel(controlbutton1)
        self.label_M.setObjectName("label_M")
        self.horizontalLayout_2.addWidget(self.label_M)

        self.verticalLayout.addLayout(self.horizontalLayout_2)

        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.verticalLayout.addLayout(self.horizontalLayout_3)
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")

        # 任务大小
        self.tasksize = QtWidgets.QLabel(controlbutton1)
        self.tasksize.setObjectName("tasksize")
        self.horizontalLayout_3.addWidget(self.tasksize)

        self.comboBox3 = QtWidgets.QComboBox(controlbutton1)  # 创建一个下拉文本框
        self.comboBox3.setObjectName("comboBox3")  # 设置对象名称
        self.comboBox3.addItem("20,33,50")  # 添加选项
        self.comboBox3.addItem("25,80,100")  # 添加选项
        # self.comboBox3.addItem("100,100,100")  # 添加选项
        # self.comboBox3.addItem("200,200,200")  # 添加选项
        self.comboBox3.addItem("20,25,33,50,80,100")  # 添加选项
        # 将下拉文本框添加到水平布局
        self.horizontalLayout_3.addWidget(self.comboBox3)

        # 任务大小单位
        self.label_M2 = QtWidgets.QLabel(controlbutton1)
        self.label_M2.setObjectName("label_M2")
        self.horizontalLayout_3.addWidget(self.label_M2)

        self.verticalLayout.addLayout(self.horizontalLayout_3)

        self.horizontalLayout_3.setStretch(0, 1)
        self.horizontalLayout_3.setStretch(1, 1)
        self.horizontalLayout_3.setStretch(2, 1)
        self.horizontalLayout_3.setStretch(3, 1)

        self.verticalLayout.setStretch(0, 1)
        self.verticalLayout.setStretch(1, 1)
        self.verticalLayout.setStretch(2, 1)

        self.horizontalLayout_5 = QtWidgets.QHBoxLayout(controlbutton1)
        self.horizontalLayout_5.setObjectName("horizontalLayout_5")
        # 将垂直布局添加到另一个水平布局。
        self.horizontalLayout_5.addLayout(self.verticalLayout)
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")

        self.runInit = QtWidgets.QPushButton(controlbutton1)
        self.runInit.setMinimumSize(QtCore.QSize(0, 50))
        self.runInit.setObjectName("runInit")
        self.verticalLayout.addWidget(self.runInit)
        self.runInit.clicked.connect(self.InitTask)

        self.horizontalLayout_5.addLayout(self.verticalLayout)
        self.horizontalLayout_5.setStretch(0, 3)
        self.horizontalLayout_5.setStretch(1, 1)
        # 调用 retranslateUi 方法，该方法通常用于设置或更新组件的文本。
        self.retranslateUi(controlbutton1)

    def retranslateUi(self, controlbutton1):
        _translate = QtCore.QCoreApplication.translate
        controlbutton1.setWindowTitle(_translate("controlbutton1", "车辆管理"))
        self.bandwidth_choose.setText(_translate("bandwidth_choose", "带宽需求：     "))
        self.fre_need.setText(_translate("fre_need", "频率需求：     "))
        self.workload.setText(_translate("workload", "工作负载：     "))
        self.tasksize.setText(_translate("tasksize", "任务大小：     "))
        self.label_Hz.setText(_translate("controlbutton", "KHz"))
        self.label_fn.setText(_translate("label_fn_choose", "MHz"))
        self.label_M.setText(_translate("controlbutton", "Mcyc"))
        self.label_M2.setText(_translate("controlbutton", "kb"))
        self.runInit.setText(_translate("controlbutton1", "确定"))

    def InitTask(self):
        global count_InitTask
        # if not self.vfrequency.text().strip():
        #     QtWidgets.QMessageBox.information(None, "提示", "频率不能为空")
        current_bandwidth = self.comboBox.currentText()
        bandwidth = current_bandwidth.split(',')

        current_fre = self.comboBox4.currentText()
        frequency = current_fre.split(',')

        current_workload = self.comboBox2.currentText()
        workload = current_workload.split(',')

        current_taskSize = self.comboBox3.currentText()
        taskSize = current_taskSize.split(',')

        desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
        excel_file_path = os.path.join(desktop_path, 'Dataset.xlsx')
        # excel_file_path = os.path.join(desktop_path, 'UNCON.xlsx')

        if os.path.exists(excel_file_path):
            df = pd.read_excel(excel_file_path)
            car_list = []
            for index, row in df.iterrows():
                row_list = list(row)
                row_list[6] = random.choice(bandwidth)
                row_list[7] = random.choice(frequency)
                row_list[8] = random.choice(workload)
                row_list[9] = random.choice(taskSize)
                # row_list[5] = float(self.vfrequency.text())
                car_list.append(row_list)
            # 创建Excel表格
            wb = Workbook()
            ws = wb.active

            ws.append(['编号', '横坐标', '纵坐标', '速度', '时刻', '是否为任务车辆', '任务带宽需求', '任务频率需求',
                         '工作负载', '任务大小', '本地时间', '本地能耗', '基站时间', '基站能耗',
                         '云卸载时间','云卸载能耗'])
            for row_data in car_list:
                ws.append(row_data)
            count_InitTask += 1
            # 保存表格
            wb.save(excel_file_path)
            QtWidgets.QMessageBox.information(None, "提示", "任务生成成功")
        else:
            QtWidgets.QMessageBox.information(None, "提示", "请先生成数据集")


class controlbutton1(QWidget, Ui_controlbutton1):
    # __init__是一个特殊的方法，称为类的构造函数。当创建这个类的新实例时，这个方法会被自动调用。
    def __init__(self):
        # 是调用父类的构造函数。在这里，它调用了QWidget和Ui_controlbutton的构造函数
        super(controlbutton1, self).__init__()
        # 调用Ui_controlbutton类中定义的setupUi方法，用于设置用户界面。这个方法需要一个参数，这里传入的是self，表示的是当前类的实例。
        self.setupUi(self)


class Ui_controlbutton2(object):
    def setupUi(self, controlbutton2):
        # self.count = count
        # print(count)
        # 设置窗口的名称和大小
        controlbutton2.setObjectName("controlbutton2")
        controlbutton2.resize(300, 300)

        # 创建一个垂直布局
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")

        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")

        # 创建一个标签，并将其添加到 controlbutton 窗口
        # 道路选择
        self.road_choose = QtWidgets.QLabel(controlbutton2)
        self.road_choose.setObjectName("road_choose")
        self.horizontalLayout.addWidget(self.road_choose)

        self.comboBox = QtWidgets.QComboBox(controlbutton2)  # 创建一个下拉文本框
        self.comboBox.setObjectName("comboBox")  # 设置对象名称
        self.comboBox.addItem("城市普通道路")  # 添加选项 "基站"
        self.comboBox.addItem("城市十字路口")  # 添加选项 "路侧单元"
        self.comboBox.addItem("高速公路")  # 添加选项 云中心"
        self.comboBox.addItem("乡村道路")
        # 将下拉文本框添加到水平布局
        self.horizontalLayout.addWidget(self.comboBox)

        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout.addItem(spacerItem)

        # 将水平布局添加到垂直布局。
        self.verticalLayout.addLayout(self.horizontalLayout)

        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")

        # 通信方式选择
        self.comm_choose = QtWidgets.QLabel(controlbutton2)
        self.comm_choose.setObjectName("comm_choose")
        self.horizontalLayout_2.addWidget(self.comm_choose)

        self.comboBox2 = QtWidgets.QComboBox(controlbutton2)  # 创建一个下拉文本框
        self.comboBox2.setObjectName("comboBox2")  # 设置对象名称
        self.comboBox2.addItem("车辆间通信")  # 添加选项 "基站"
        self.comboBox2.addItem("车辆与基站通信")  # 添加选项 "路侧单元"
        self.comboBox2.addItem("车辆与云通信")  # 添加选项 云中心"
        self.comboBox2.addItem("车辆与基站/云通信")  # 添加选项 云中心"
        # 将下拉文本框添加到水平布局
        self.horizontalLayout_2.addWidget(self.comboBox2)
        spacerItem2 = QtWidgets.QSpacerItem(80, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_2.addItem(spacerItem2)

        self.verticalLayout.addLayout(self.horizontalLayout_2)

        self.horizontalLayout_2.setStretch(0, 1)
        self.horizontalLayout_2.setStretch(1, 1)

        self.verticalLayout.setStretch(0, 1)
        self.verticalLayout.setStretch(1, 1)

        self.horizontalLayout_4 = QtWidgets.QHBoxLayout(controlbutton2)
        self.horizontalLayout_4.setObjectName("horizontalLayout_4")
        # 将垂直布局添加到另一个水平布局。
        self.horizontalLayout_4.addLayout(self.verticalLayout)
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")

        self.road_fix = QtWidgets.QPushButton(controlbutton2)
        self.road_fix.setMinimumSize(QtCore.QSize(0, 50))
        self.road_fix.setObjectName("run2")
        self.verticalLayout.addWidget(self.road_fix)
        # self.road_fix.clicked.connect(self.onButtonClicked)  # 将按钮点击事件与函数绑定

        self.horizontalLayout_4.addLayout(self.verticalLayout)
        self.horizontalLayout_4.setStretch(0, 3)
        self.horizontalLayout_4.setStretch(1, 1)
        # 调用 retranslateUi 方法，该方法通常用于设置或更新组件的文本。
        self.retranslateUi(controlbutton2)

    def retranslateUi(self, controlbutton2):
        _translate = QtCore.QCoreApplication.translate
        controlbutton2.setWindowTitle(_translate("controlbutton2", "道路通信管理"))
        self.road_choose.setText(_translate("road_choose", "道路选择：  "))
        self.comm_choose.setText(_translate("comm_choose", "通信方式："))
        self.road_fix.setText(_translate("controlbutton", "确定/取消"))


class controlbutton2(QWidget, Ui_controlbutton2):
    # __init__是一个特殊的方法，称为类的构造函数。当创建这个类的新实例时，这个方法会被自动调用。
    def __init__(self):
        # 是调用父类的构造函数。在这里，它调用了QWidget和Ui_controlbutton的构造函数
        super(controlbutton2, self).__init__()
        # 调用Ui_controlbutton类中定义的setupUi方法，用于设置用户界面。这个方法需要一个参数，这里传入的是self，表示的是当前类的实例。
        self.setupUi(self)


class textwidget(QtWidgets.QWidget):
    def __init__(self):
        # 调用父类 QtWidgets.QWidget 的构造函数，以初始化 textwidget 类的实
        super(textwidget, self).__init__()
        # 这行代码设置了 textwidget 实例的对象名称为 "textwidget"
        self.setObjectName("textwidget")
        # 设置了 textwidget 实例的大小为 400x400 像素
        self.resize(400, 400)
        # 创建了一个新的垂直布局，并将其设置为 textwidget 实例的布局
        self.verticalLayout = QtWidgets.QVBoxLayout(self)
        self.verticalLayout.setObjectName("verticalLayout")
        # 创建了一个新的文本编辑器。
        self.textEdit = QtWidgets.QTextEdit(self)
        self.textEdit.setObjectName("textEdit")
        # 将文本编辑器添加到垂直布局中。
        self.verticalLayout.addWidget(self.textEdit)
        # 调用 retranslateUi 方法，该方法通常用于设置或更新用户界面的文本内容
        self.retranslateUi()
        # 这行代码将所有的 Qt 信号和槽连接起来。这是 Qt 的元对象系统的一部分，它允许在运行时进行对象间的通信
        QtCore.QMetaObject.connectSlotsByName(self)

    # 设置窗口标题的方法
    def retranslateUi(self):
        # retranslateUi 是一个方法，它的主要目的是为了在更改应用程序的语言时，能够动态地更改用户界面的文本
        _translate = QtCore.QCoreApplication.translate
        # 设置窗口的标题。_translate("textwidget", "Form") 会根据当前的翻译上下文，将 "Form" 翻译成相应的语言，
        # 然后设置为窗口的标题。如果没有找到对应的翻译，就会直接使用 "Form" 作为窗口的标题。
        self.setWindowTitle(_translate("textwidget", "Form"))


# 定义了一个名为Ui_MainWindow的类，该类用于设置和配置主窗口的各种组件和属性
class Ui_MainWindow(object):
    tpFlag = 0
    agFlag = 0

    # 一个方法，用于设置和配置主窗口的各种组件和属性
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1000, 900)  # 宽 高
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.verticalLayout_3 = QtWidgets.QVBoxLayout(self.centralwidget)
        self.verticalLayout_3.setObjectName("verticalLayout_3")

        self.groupBox = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox.setObjectName("groupBox")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.groupBox)
        self.verticalLayout.setObjectName("verticalLayout")
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")

        # spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        # self.horizontalLayout.addItem(spacerItem)

        self.label = QtWidgets.QLabel(self.groupBox)
        self.label.setObjectName("label")
        self.horizontalLayout.addWidget(self.label)

        self.userlineEdit = QtWidgets.QLineEdit(self.groupBox)
        self.userlineEdit.setMinimumSize(QtCore.QSize(0, 30))
        self.userlineEdit.setObjectName("userlineEdit")
        self.horizontalLayout.addWidget(self.userlineEdit)

        self.label_2 = QtWidgets.QLabel(self.groupBox)
        self.label_2.setObjectName("label_2")
        self.horizontalLayout.addWidget(self.label_2)

        spacerItem2 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout.addItem(spacerItem2)

        self.runButton = QtWidgets.QPushButton(self.groupBox)
        self.runButton.setMinimumSize(QtCore.QSize(0, 50))
        self.runButton.setObjectName("runButton")
        self.horizontalLayout.addWidget(self.runButton)

        spacerItem3 = QtWidgets.QSpacerItem(60, 60, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout.addItem(spacerItem3)

        self.stopButton = QtWidgets.QPushButton(self.groupBox)
        self.stopButton.setMinimumSize(QtCore.QSize(0, 50))
        self.stopButton.setObjectName("stopButton")
        self.horizontalLayout.addWidget(self.stopButton)

        spacerItem4 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout.addItem(spacerItem4)

        self.openButton = QtWidgets.QPushButton(self.groupBox)
        self.openButton.setMinimumSize(QtCore.QSize(0, 50))
        self.openButton.setObjectName("openButton")
        self.horizontalLayout.addWidget(self.openButton)

        spacerItem5 = QtWidgets.QSpacerItem(60, 60, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout.addItem(spacerItem5)

        self.horizontalLayout.setStretch(0, 1)
        self.horizontalLayout.setStretch(1, 10)
        self.horizontalLayout.setStretch(2, 1)
        self.horizontalLayout.setStretch(3, 1)
        self.horizontalLayout.setStretch(4, 1)
        self.horizontalLayout.setStretch(5, 1)

        self.verticalLayout.addLayout(self.horizontalLayout)
        self.verticalLayout_3.addWidget(self.groupBox)

        self.groupBox_2 = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox_2.setObjectName("groupBox_2")
        self.horizontalLayout_4 = QtWidgets.QHBoxLayout(self.groupBox_2)
        self.horizontalLayout_4.setObjectName("horizontalLayout_4")
        spacerItem6 = QtWidgets.QSpacerItem(20, 10, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_4.addItem(spacerItem6)

        self.templateButton = QtWidgets.QPushButton(self.groupBox_2)
        self.templateButton.setMinimumSize(QtCore.QSize(0, 50))
        self.templateButton.setObjectName("templateButton")
        self.horizontalLayout_4.addWidget(self.templateButton)
        spacerItem7 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_4.addItem(spacerItem7)

        self.upAlButton = QtWidgets.QPushButton(self.groupBox_2)
        self.upAlButton.setMinimumSize(QtCore.QSize(0, 50))
        self.upAlButton.setObjectName("upAlButton")
        self.horizontalLayout_4.addWidget(self.upAlButton)

        spacerItem8 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_4.addItem(spacerItem8)

        self.runAlButton = QtWidgets.QPushButton(self.groupBox_2)
        self.runAlButton.setMinimumSize(QtCore.QSize(0, 50))
        self.runAlButton.setObjectName("runAlButton")
        self.horizontalLayout_4.addWidget(self.runAlButton)

        spacerItem9 = QtWidgets.QSpacerItem(50, 50, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_4.addItem(spacerItem9)

        self.IntroButton = QtWidgets.QPushButton(self.groupBox_2)
        self.IntroButton.setMinimumSize(QtCore.QSize(0, 50))
        self.IntroButton.setObjectName("IntroButton")
        self.horizontalLayout_4.addWidget(self.IntroButton)

        spacerItem10 = QtWidgets.QSpacerItem(60, 60, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.horizontalLayout_4.addItem(spacerItem10)

        self.horizontalLayout_4.setStretch(0, 1)
        self.horizontalLayout_4.setStretch(1, 1)
        self.horizontalLayout_4.setStretch(2, 1)
        self.horizontalLayout_4.setStretch(3, 1)
        self.horizontalLayout_4.setStretch(4, 1)

        self.verticalLayout_3.addWidget(self.groupBox_2)

        self.groupBox_3 = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox_3.setObjectName("groupBox_3")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(self.groupBox_3)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        # self.tableWidget = QtWidgets.QTableWidget(self.groupBox_3)

        image = QtGui.QImage("images/Intersection.png")  # 加载图像
        label = QtWidgets.QLabel()
        label.setPixmap(QtGui.QPixmap.fromImage(image))
        label.setScaledContents(True)  # 拉伸图像以填满 QLabel 的大小

        # 设置 QLabel 的宽度和高度
        label.setFixedWidth(600)  # 设置宽度为 200 像素
        label.setFixedHeight(600)  # 设置高度为 200 像素

        # 添加 QLabel 到垂直布局
        self.verticalLayout_2.addWidget(label)

        self.verticalLayout_3.addWidget(self.groupBox_3)
        # self.verticalLayout_3.setStretch(0, 50)

        self.verticalLayout_3.setStretch(0, 1)
        self.verticalLayout_3.setStretch(1, 1)
        self.verticalLayout_3.setStretch(2, 10)

        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        # 创建了三个停靠窗口（DockWidget），并将它们添加到主窗口中。
        # 创建QDockWidget窗口（标题，自身窗口）
        self.items = QDockWidget('道路通信设置', self)
        self.items.setMinimumSize(450, 150)  # 设置最小宽度和高度
        self.items.setMaximumSize(450, 200)  # 设置最大宽度和高度
        self.items.resize(450, 200)  # 设置停靠窗口的宽度和高度
        self.items2 = QDockWidget('车辆任务设置', self)
        self.items2.setMinimumSize(450, 150)  # 设置最小宽度和高度
        self.items2.setMaximumSize(450, 275)  # 设置最大宽度和高度
        self.items2.resize(300, 200)  # 设置停靠窗口的宽度和高度为 300 像素
        self.items3 = QDockWidget('边缘资源设置', self)
        self.items3.setMinimumSize(450, 200)  # 设置最小宽度和高度
        self.items3.setMaximumSize(450, 400)  # 设置最大宽度和高度
        self.items3.resize(450, 300)  # 设置停靠窗口的宽度和高度为 300 像素
        self.items4 = QDockWidget('仿真结果', self)
        self.items4.setMinimumSize(450, 200)  # 设置最小宽度和高度
        self.items4.setMaximumSize(450, 400)  # 设置最大宽度和高度
        self.items4.resize(300, 300)  # 设置停靠窗口的宽度和高度为 300 像素

        # 创建了三个自定义的QWidget对象
        self.Widget = controlbutton2()
        self.Widget2 = controlbutton1()
        self.Widget3 = controlbutton()
        self.Widget4 = textwidget()

        # 将这些QWidget对象设置为对应的DockWidget的内容。
        # 在窗口区域设置QWidget，添加列表控件
        self.items.setWidget(self.Widget)
        self.items2.setWidget(self.Widget2)
        self.items3.setWidget(self.Widget3)
        self.items4.setWidget(self.Widget4)
        # 设置dock窗口是否可以浮动，True，运行浮动在外面，自动与主界面脱离，False，默认浮动主窗口内，可以手动脱离
        self.items.setFloating(False)
        self.items2.setFloating(False)
        self.items3.setFloating(False)
        self.items4.setFloating(False)
        # 将这些DockWidget添加到主窗口的左侧区域。
        self.addDockWidget(Qt.LeftDockWidgetArea, self.items)
        self.addDockWidget(Qt.LeftDockWidgetArea, self.items2)
        self.addDockWidget(Qt.LeftDockWidgetArea, self.items3)
        self.addDockWidget(Qt.LeftDockWidgetArea, self.items4)

        # self.server_thread = ServerThread(self.Widget4)
        # self.server_thread.start()

        self.retranslateUi(MainWindow)
        # 将各种按钮的点击事件连接到相应的处理函数。、
        # self.showtablepushButton.clicked.connect(self.showtablesname)
        # self.selectpushButton.clicked.connect(self.showtable)
        # self.deletepushButton.clicked.connect(self.deletetable)
        # self.runButton.clicked.connect(self.run)
        self.Widget.road_fix.clicked.connect(self.onButtonClicked)  # 将按钮点击事件与函数绑定
        self.runButton.clicked.connect(self.run)
        self.stopButton.clicked.connect(self.stop)
        self.openButton.clicked.connect(self.openFile)
        self.templateButton.clicked.connect(self.teAl)
        self.upAlButton.clicked.connect(self.upAlgorithm)
        self.runAlButton.clicked.connect(self.runAlgorithm)
        self.IntroButton.clicked.connect(self.Introduction)

        QtCore.QMetaObject.connectSlotsByName(MainWindow)


        # self.start_udp_thread()
        # # 创建一个新的线程用于接收数据
        # self.thread = threading.Thread(target=self.receive_data)
        # self.thread.start()

    def start_udp_thread(self):
        self.Widget4.textEdit.clear()
        self.udp_thread = threading.Thread(target=self.receive_udp)
        self.udp_thread.start()

    def receive_udp(self):
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.bind(('localhost', 12345))
        while True:
            data, addr = sock.recvfrom(1024)
            self.Widget4.textEdit.append(data.decode())

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "车辆边缘计算仿真平台"))
        # self.label_6.setText(_translate("MainWindow", "调试模式开关   "))
        self.groupBox.setTitle(_translate("MainWindow", "运行模块"))
        self.label.setText(_translate("MainWindow", "运行时间："))
        self.userlineEdit.setText(_translate("MainWindow", "2500"))
        self.label_2.setText(_translate("MainWindow", "s      "))

        self.runButton.setText(_translate("MainWindow", "运行"))
        self.stopButton.setText(_translate("MainWindow", "退出程序"))
        self.openButton.setText(_translate("MainWindow", "打开文件"))

        self.groupBox_2.setTitle(_translate("MainWindow", "算法模块"))
        self.templateButton.setText(_translate("MainWindow", "算法模板"))
        self.upAlButton.setText(_translate("MainWindow", "上传算法"))
        self.runAlButton.setText(_translate("MainWindow", "运行算法"))
        self.IntroButton.setText(_translate("MainWindow", "平台介绍"))
        self.groupBox_3.setTitle(_translate("MainWindow", "道路信息显示"))

    def onButtonClicked(self):
        global count
        global selected_road_text
        global selected_comm_text
        count += 1
        if (count % 2 == 1):
            self.Widget.comboBox.setEnabled(False)  # 按下按钮后禁用下拉文本框
            self.Widget.comboBox2.setEnabled(False)  # 按下按钮后禁用下拉文本框
            selected_road_text = self.Widget.comboBox.currentText()  # 获取下拉框当前选择的文本
            selected_comm_text = self.Widget.comboBox2.currentText()  # 获取下拉框当前选择的文本
            Ui_MainWindow.changeImage(self, selected_road_text)
        else:
            self.Widget.comboBox.setEnabled(True)  # 按下按钮后禁用下拉文本框
            self.Widget.comboBox2.setEnabled(True)  # 按下按钮后禁用下拉文本框

    def changeImage(self, selected_text):
        try:
            if count % 2 == 1:
                label = QtWidgets.QLabel()
                if selected_text == "城市普通道路":
                    self.image = QtGui.QImage("images/coordinate/urban.png")
                    label.setPixmap(QtGui.QPixmap.fromImage(self.image))
                    label.setScaledContents(True)  # 拉伸图像以填满 QLabel 的大小
                    # 设置 QLabel 的宽度和高度
                    label.setFixedWidth(700)  # 设置宽度为 200 像素
                    label.setFixedHeight(450)  # 设置高度为 200 像素
                elif selected_text == "城市十字路口":
                    self.image = QtGui.QImage("images/coordinate/intersection.png")
                    label.setPixmap(QtGui.QPixmap.fromImage(self.image))
                    label.setScaledContents(True)  # 拉伸图像以填满 QLabel 的大小
                    # 设置 QLabel 的宽度和高度
                    label.setFixedWidth(800)  # 设置宽度为 200 像素
                    label.setFixedHeight(500)  # 设置高度为 200 像素
                elif selected_text == "高速公路":
                    self.image = QtGui.QImage("images/coordinate/express.png")
                    label.setPixmap(QtGui.QPixmap.fromImage(self.image))
                    label.setScaledContents(True)  # 拉伸图像以填满 QLabel 的大小
                    # 设置 QLabel 的宽度和高度
                    label.setFixedWidth(800)  # 设置宽度为 200 像素
                    label.setFixedHeight(500)  # 设置高度为 200 像素
                else:
                    self.image = QtGui.QImage("images/coordinate/rural.png")
                    label.setPixmap(QtGui.QPixmap.fromImage(self.image))
                    label.setScaledContents(True)  # 拉伸图像以填满 QLabel 的大小
                    # 设置 QLabel 的宽度和高度
                    label.setFixedWidth(600)  # 设置宽度为 200 像素
                    label.setFixedHeight(300)  # 设置高度为 200 像素
            # 移除旧的 QLabel
            for i in reversed(range(self.verticalLayout_2.count())):
                self.verticalLayout_2.itemAt(i).widget().setParent(None)
            # 添加 QLabel 到垂直布局
            self.verticalLayout_2.addWidget(label)
        except Exception as e:
            print("An error occurred:", e)

    def run(self):
        self.start_udp_thread()

        run_time = self.userlineEdit.text()
        if count % 2 == 1:
            # 获取当前脚本所在的目录
            script_directory = os.path.dirname(os.path.abspath(__file__))
            if selected_road_text == "城市普通道路":
                # 在脚本目录下构建一个 Python 脚本的完整路径
                python_script_path = os.path.join(script_directory, 'simu_Urban_ordinary.py')
                # 使用 subprocess.run 运行 Python 脚本
                # subprocess.run(["python", python_script_path, str(run_time)], check=True, shell=True)
                # subprocess.run(["python", "simu_Urban_ordinary.py", str(run_time)], check=True, shell=True)
                # subprocess.run(["python", "2.py", str(run_time)], check=True, shell=True)
            elif selected_road_text == "城市十字路口":
                # subprocess.run(["python", "2.py", str(run_time)], check=True, shell=True)
                # subprocess.run(["python", "3.py"], check=True, shell=True)
                # subprocess.run(["python", "3.py", str(run_time)], check=True, shell=True)
                python_script_path = os.path.join(script_directory, 'simu_Intersection.py')
                # subprocess.run(["python", "simu_Intersection.py", str(run_time)], check=True, shell=True)
            elif selected_road_text == "高速公路":
                python_script_path = os.path.join(script_directory, 'simu_Expressway.py')
                # subprocess.run(["python", "simu_Expressway.py", str(run_time)], check=True, shell=True)
            elif selected_road_text == "乡村道路":
                python_script_path = os.path.join(script_directory, 'simu_Rural_road.py')
                # subprocess.run(["python", "simu_Rural_road.py", str(run_time)], check=True, shell=True)
            # 使用 subprocess.run 运行 Python 脚本
            subprocess.run(["python", python_script_path, str(run_time)], check=True, shell=True)
        else:
            QtWidgets.QMessageBox.information(self.centralwidget, "提示", "请确定道路")


    def stop(self):
        quit()

    def openFile(self):
        desktop_path = os.path.expanduser("~/Desktop")  # 获取桌面路径
        file_path = os.path.join(desktop_path, "Dataset.xlsx")  # 创建文件路径
        # file_path = os.path.join(desktop_path, "UNCON.xlsx")  # 创建文件路径
        if os.path.exists(file_path):
            os.startfile(file_path)
        else:
            QtWidgets.QMessageBox.information(self.centralwidget, "提示", "没有数据集文件，请先生成数据集")

    def teAl(self):
        self.example_ag = algorithm_example()
        self.example_ag.show()

    def Introduction(self):
        self.example_ag = Introduction_example()
        self.example_ag.show()

    def upAlgorithm(self):
        # 检查当前目录是否存在 Al.py 文件，如果存在则删除
        if os.path.exists('Al.py'):
            os.remove('Al.py')

        # 创建一个新的空白 Al.py 文件
        open('Al.py', 'a').close()

        root = tk.Tk()
        root.withdraw()  # 隐藏主窗口

        # 弹出文件对话框，让用户选择文件
        selected_algorithm_file = filedialog.askopenfilename(initialdir=os.getcwd(), title="选择算法文件")
        if selected_algorithm_file:
            shutil.copy(selected_algorithm_file, 'Al.py')

    def runAlgorithm(self):
        global external_car_fre
        # print(external_car_fre)
        global external_base_x
        global external_base_y
        global external_base_band
        global external_base_fre
        global external_base_cov

        global external_cloud_x
        global external_cloud_y
        global external_cloud_band
        global external_cloud_fre
        global external_cloud_cov

        global task_loss
        global SC
        global count
        global selected_comm_text
        try:
            with open('Al.py', 'r', encoding='utf-8') as file:
                al_code = file.read()

            # 执行 Al.py，并传递参数
            exec(al_code, globals(), locals())

            # 根据条件选择执行的函数
            if (count % 2 == 0):
                QtWidgets.QMessageBox.information(self.centralwidget, "提示", "请确定通信方式")

                # 根据 selected_text 执行不同的函数
            else:
                if selected_comm_text == "车辆与基站/云通信":
                    task_loss, SC = vehicle_to_base_cloud(0, 0, external_car_fre, external_base_x, external_base_y,
                                          external_base_cov, external_base_band, external_base_fre,
                                          external_cloud_x, external_cloud_y, external_cloud_cov,
                                          external_cloud_band, external_cloud_fre)

                elif selected_comm_text == "车辆与云通信":
                    task_loss, SC = vehicle_to_cloud(0, 0, external_base_x, external_cloud_y, external_cloud_cov, external_cloud_band, external_cloud_fre)
                elif selected_comm_text == "车辆与基站通信":
                    task_loss, SC = vehicle_to_base(0, 0, external_base_x, external_base_y, external_base_cov, external_base_band, external_base_fre)
                elif selected_comm_text == "车辆间通信":
                    task_loss, SC = vehicle_to_vehicle(0, 0, external_car_fre)
                task_loss_str = str(task_loss)
                SC_str = str(SC)
                QMessageBox.information(self.centralwidget, "运行完成",
                                        f"任务丢失数量为 {task_loss_str}，系统成本为 {SC_str}")
        except FileNotFoundError:
            print("Al.py 文件未找到")
            QtWidgets.QMessageBox.information(self.centralwidget, "错误", "缺少文件")
        except Exception as e:
            if "not supported between instances of 'int' and 'NoneType'" in str(e):
                QtWidgets.QMessageBox.information(self.centralwidget, "错误", "请进行资源设置")
            else:
                print(f"执行 Al.py 文件时发生错误: {e}")
                QtWidgets.QMessageBox.information(self.centralwidget, "错误", "执行算法时发生错误: {e}")

    # 用来更新运行信息的。它将传入的文本参数text添加到Widget2的文本编辑器中。
    def updateRunInfo(self, text):
        self.Widget4.textEdit.append(1)
        # self.Widget4.textEdit.append(text)

    def getResult(self, text):
        pid = os.getpid()
        if text == "0":
            QtWidgets.QMessageBox.information(self.centralwidget, "提示",
                                              "程序编写出错，请检查导入拓扑所用端口是否能正常使用，导入算法是否具有语法问题，以及系统是否具有算法所需的python模块")
            run('taskkill /F /im python.exe /FI ' + '"PID ne ' + str(pid) + '"', shell=True)
        else:
            self.Widget3.textEdit.setPlainText(text)
            run('taskkill /F /im python.exe /FI ' + '"PID ne ' + str(pid) + '"', shell=True)


# 这段代码定义了一个名为 MainWindow 的类，该类继承自 QMainWindow 和 Ui_MainWindow。QMainWindow 是 PyQt5 中的一个类，
# 用于创建主窗口。Ui_MainWindow 是用户自定义的类，通常用于设置主窗口的用户界面。
class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.setupUi(self)


class algorithm_example(QtWidgets.QWidget):
    def __init__(self):
        super(algorithm_example, self).__init__()
        self.setObjectName("example_ag")
        self.resize(1000, 800)
        self.setWindowTitle("算法模板格式")
        self.textEdit = QtWidgets.QTextEdit(self)
        self.textEdit.setEnabled(True)
        self.textEdit.setGeometry(QtCore.QRect(30, 30, 970, 700))
        self.textEdit.setObjectName("textEdit")
        _translate = QtCore.QCoreApplication.translate
        self.textEdit.setHtml(_translate("showAlgorithm",
                                         "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
                                         "<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
                                         "p, li { white-space: pre-wrap; }\n"
                                         "</style></head><body style=\" font-family:\'SimSun\'; font-size:9pt; font-weight:400; font-style:normal;\">\n"
                                         "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; background-color:#ffffff;\">"
                                         "<span style=\" font-family:\'Menlo\'; font-style:italic; color:#808080;\">"
                                         "# import需求模块<br /><br /># 用户自定义函数区<br /><br />"
                                         "# 系统成本<br />SC = 0<br /># 任务丢失<br />task_loss = 0<br /><br />"
                                         "# 车辆间卸载任务函数,vfrequency为传入的车辆频率<br /></span>"
                                         "<span style=\" font-family:\'Menlo\'; font-weight:600; color:#000080;\">def </span><span style=\" font-family:\'Menlo\'; color:#000000;\">vehicle_to_vehicle(task_loss,SC,vfrequency):<br /> </span>"
                                         "    <span style=\" font-family:\'Menlo\'; font-weight:600; color:#008000;\">&quot;&quot;&quot;<br />    用户编码区域<br />    &quot;&quot;&quot;<br />"
                                         "    <span style=\" font-family:\'Menlo\'; font-weight:600; color:#000080;\">return </span><span style=\" font-family:\'Menlo\'; color:#000000;\">task_loss,SC</span><br />"
                                         "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; background-color:#ffffff;\">"
                                         "<span style=\" font-family:\'Menlo\'; font-style:italic; color:#808080;\">"
                                         "# 基站卸载任务函数,bx为基站横坐标,by为基站纵坐标,bcoverage为基站覆盖范围,bbandwidth为传入的基站带宽,bfrequency为传入的基站计算频率<br /></span>"
                                         "<span style=\" font-family:\'Menlo\'; font-weight:600; color:#000080;\">def </span><span style=\" font-family:\'Menlo\'; color:#000000;\">vehicle_to_base(task_loss,SC,bx,by,bcoverage,bbandwidth,bfrequency):<br /> </span>"
                                         "    <span style=\" font-family:\'Menlo\'; font-weight:600; color:#008000;\">&quot;&quot;&quot;<br />    用户编码区域<br />    &quot;&quot;&quot;<br />"
                                         "    <span style=\" font-family:\'Menlo\'; font-weight:600; color:#000080;\">return </span><span style=\" font-family:\'Menlo\'; color:#000000;\">task_loss,SC</span>"
                                         "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; background-color:#ffffff;\">"
                                         "<span style=\" font-family:\'Menlo\'; font-style:italic; color:#808080;\">"
                                         "# 云卸载任务函数,cx为云横坐标,cy为云纵坐标,ccoverage为云覆盖范围,cbandwidth为传入的云带宽,cfrequency为传入的云计算频率<br /></span>"
                                         "<span style=\" font-family:\'Menlo\'; font-weight:600; color:#000080;\">def </span><span style=\" font-family:\'Menlo\'; color:#000000;\">vehicle_to_cloud(task_loss,SC,cx,cy,ccoverage,cbandwidth,cfrequency):<br /> </span>"
                                         "    <span style=\" font-family:\'Menlo\'; font-weight:600; color:#008000;\">&quot;&quot;&quot;<br />    用户编码区域<br />    &quot;&quot;&quot;<br />"
                                         "    <span style=\" font-family:\'Menlo\'; font-weight:600; color:#000080;\">return </span><span style=\" font-family:\'Menlo\'; color:#000000;\">task_loss,SC</span>"
                                         "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; background-color:#ffffff;\">"
                                         "<span style=\" font-family:\'Menlo\'; font-style:italic; color:#808080;\">"
                                         "# 基站/云卸载任务函数<br /></span>"
                                         "<span style=\" font-family:\'Menlo\'; font-weight:600; color:#000080;\">def </span><span style=\" font-family:\'Menlo\'; color:#000000;\">vehicle_to_base_cloud(task_loss,SC,vfrequency,bx,by,bcoverage,bbandwidth,bfrequency,cx,cy,ccoverage,cbandwidth,cfrequency):<br /> </span>"
                                         "    <span style=\" font-family:\'Menlo\'; font-weight:600; color:#008000;\">&quot;&quot;&quot;<br />    用户编码区域<br />    &quot;&quot;&quot;<br />"
                                         "    <span style=\" font-family:\'Menlo\'; font-weight:600; color:#000080;\">return </span><span style=\" font-family:\'Menlo\'; color:#000000;\">task_loss,SC</span>"
                                         "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; background-color:#ffffff;\">"
                                         "</p></body></html>"))

class Introduction_example(QtWidgets.QWidget):
    def __init__(self):
        super(Introduction_example, self).__init__()
        self.setObjectName("example_ag")
        self.resize(500, 500)
        self.setWindowTitle("平台功能介绍")
        self.textEdit = QtWidgets.QTextEdit(self)
        self.textEdit.setEnabled(True)
        self.textEdit.setGeometry(QtCore.QRect(0, 0, 500, 500))
        self.textEdit.setObjectName("textEdit")
        _translate = QtCore.QCoreApplication.translate
        self.textEdit.setHtml(_translate("showAlgorithm",
                                         "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
                                         "<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
                                         "p, li { white-space: pre-wrap; }\n"
                                         "</style></head><body style=\" font-family:\'SimSun\'; font-size:9pt; font-weight:400; font-style:normal;\">\n"
                                         "<p style=\" margin-top:12px; margin-bottom:12px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px; background-color:#ffffff;\">"
                                         "<span style=\" font-family:\'Menlo\'; font-weight:600; color:#000080;\">车辆边缘计算仿真平台 <br /><br />"
                                         "该平台的主要功能是模拟车辆行驶并生成数据集，用户可以选择上传车辆边缘计算算法进行卸载任务</span>"
                                         "</p></body></html>"))